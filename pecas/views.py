import mimetypes
from django.http import StreamingHttpResponse
from wsgiref.util import FileWrapper
from django.views.generic import View

from django.shortcuts import render,redirect
from django.http import HttpResponse,HttpResponseRedirect,JsonResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required

#from . import dados_informados,embargos_omissao,funcoes_banco,funcoes_gerais,funcoes_teses,peticao_des_dev_hon_dup,peticao_teste,teses,peticao_dev_hp_improcedente,peticoes
from . import dados_informados,embargos_omissao,funcoes_banco,funcoes_gerais,funcoes_teses,teses,peticoes

from .classes import Pasta
from openpyxl import Workbook


from .models import Estado
import pyodbc as p
import os
import json
import mysql.connector


def download_file(request,docmto):
   the_file = os.environ.get('DIR_DOCUMENTOS')+'/'+docmto
   filename = os.path.basename(the_file)
   chunk_size = 8192
   response = StreamingHttpResponse(FileWrapper(open(the_file, 'rb'), chunk_size),
                           content_type=mimetypes.guess_type(the_file)[0])
   response['Content-Length'] = os.path.getsize(the_file)
   response['Content-Disposition'] = "attachment; filename=%s" % filename
   return response


def v003_dados_da_pasta(request):

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if is_ajax:
        cod_pasta_saj = request.GET.get('opcao')

        dados_view = funcoes_banco.f001_sql(cod_pasta_saj,1)
        pasta=dados_view['pasta']
        cod_cliente=dados_view['cod_cliente']
        autor=dados_view['autor']
        comarca=dados_view['comarca']
        uf=dados_view['uf']
        orgao=dados_view['orgao']
        num_orgao=dados_view['num_orgao']
        juizo=dados_view['juizo']
        secao=dados_view['secao']
        nr_processo=dados_view['nr_processo']
        empresa=dados_view['reu']
        oabjb=dados_view['oabJB']
        publicando_nome=dados_view['publicando_nome']
        publicando_oab=dados_view['publicando_oab']
        publicando_sexo=dados_view['publicando_sexo']
        conveniado_nome=dados_view['conveniado_nome']
        conveniado_oab=dados_view['conveniado_oab']

        data=[]
        data.append({'key':0,'value':pasta})
        data.append({'key':1,'value':cod_cliente})
        data.append({'key':2,'value':autor})
        data.append({'key':3,'value':comarca})
        data.append({'key':4,'value':uf})
        data.append({'key':5,'value':orgao})
        data.append({'key':6,'value':num_orgao})
        data.append({'key':7,'value':juizo})
        data.append({'key':8,'value':secao})

        data.append({'key':9,'value':nr_processo})
        data.append({'key':10,'value':empresa})
        data.append({'key':11,'value':oabjb})
        data.append({'key':12,'value':publicando_nome})
        data.append({'key':13,'value':publicando_oab})
        data.append({'key':14,'value':publicando_sexo})
        data.append({'key':15,'value':conveniado_nome})
        data.append({'key':16,'value':conveniado_oab})

        data = json.dumps(data)
    else:
        data=[]
    return HttpResponse(data, content_type='application/json')

#------------------------------------------------------------------------

@login_required
def v007_embargos(request):
    if (request.method == "POST"):
        #request.session['pasta']=request.POST['pasta']
        data_publicacao=request.POST['data_publicacao']
        if data_publicacao =='':
            data_publicacao='2021-01-01'
        #data_publicacao=funcoes_gerais.convert_data_br(data_publicacao)

        hanulidadepublic=request.POST['hanulidadepublic']
        motivo_embargos=request.POST['motivo_embargos']
        pasta = request.POST['pasta']

        if hanulidadepublic=='S':
            request.session['nulidade'] = 'S'
        else:
            request.session['nulidade'] = 'N'

        if motivo_embargos=='OMISSAO':
            return redirect('pecas:embargos-omissao', pasta=pasta, hanulidadepublic = hanulidadepublic,data_publicacao = data_publicacao)
        elif motivo_embargos=='ULTRA PETITA-CE':
            return redirect('pecas:embargos-ultrapetita')
        elif motivo_embargos=='CONTRADICAO':
            return redirect('pecas:embargos-contradicao')
    else:
        dados = {
            'tipo_da_peca': 'Embargos'
        }
        return render(request, 'pecas/embargos.html', dados)



#@login_required
def v008_embargos_omissao(request,pasta,hanulidadepublic,data_publicacao):
    if (request.method == "POST"):
        pasta = request.POST['pasta']
        #hanulidadepublic = request.POST['hanulidadepublic']
        data_publicacao = request.POST['data_publicacao']

        teses_emb_omissao = teses.teses_embargos_omissao

        respostas = {
            'intimacao_mp':request.POST['intimacao_mp'],
            'cj_peremp_litisp':request.POST['cj_peremp_litisp'],
            'prescricao':request.POST['prescricao'],
            'pagamento_adm':request.POST['pagamento_adm'],
            'prop_inadimp':request.POST['prop_inadimp'],
            'prop_inadimp_com_pagto_Adm':request.POST['prop_inadimp_com_pagto_Adm'],
            'lesao_pre':request.POST['lesao_pre'],
            'omissao_regulacao_8':request.POST['omissao_regulacao_8'],
            'juros_citacao':request.POST['juros_citacao'],
            'correcao_monetaria':request.POST['correcao_monetaria']
        }

        if hanulidadepublic=='S':
            teses_emb_omissao['Nulidade']='S'


        teses_emb_omissao = funcoes_teses.f002_teses_embargos_omissao(respostas,teses_emb_omissao)

        dados_complementares = dados_informados.dados_embargos_omissao()

        dados_complementares['data_public']=data_publicacao
        dados_complementares['valor_pagamento_adm']=request.POST['valor_pagamento_adm']
        valor = request.POST['valor_pagamento_adm']
        dados_complementares['num_processo_vinculado']=request.POST['num_processo_vinculado']
        dados_complementares['juizo_vinculado']=request.POST['juizo_vinculado']
        dados_complementares['juizo_lpe']=request.POST['juizo_lpe']
        dados_complementares['num_proc_lpe']=request.POST['num_proc_lpe']
        dados_complementares['data_lpe']=request.POST['data_lpe']
        dados_complementares['desc_lpe']=request.POST['desc_lpe']
        dados_complementares['local_diligencia']=request.POST['local_diligencia']

        dados_pasta = funcoes_banco.f002_sql(pasta)

        dados_complementares['adv_publicando'] = dados_pasta['publicando_nome']
        dados_complementares['nr_processo'] = dados_pasta['nr_processo']
        dados_complementares['cod_cliente'] = dados_pasta['cod_cliente']


        valor_extenso=funcoes_gerais.valor_por_extenso(valor)

        if valor!="":
            dados_complementares['pagamento_adm']="R$ "+valor+" ("+valor_extenso+")"
        else:
            dados_complementares['pagamento_adm']=''

        #return HttpResponse("<h1>Teste</h1>")


        file = embargos_omissao.peticoes(teses=teses_emb_omissao,dados_compl=dados_complementares,dados_pasta=dados_pasta)

        #file = templates.embargosOmissao(context=dados_complementares)


        #return redirect('pecas:templatepeca', docmto=file)

        return redirect('pecas:download', docmto=file)
    else:

        #pasta = request.session.get('pasta')
        tese_nulidade = 'S' #request.session.get('tese_nulidade')
        dados = funcoes_banco.f001_sql(pasta,1)

        autor='nome do autor'
        dados={
            'pasta':pasta,
            'autor':dados['autor'],
            'comarca':dados['comarca'],
            'uf':dados['uf'],
            'cod_cliente':dados['cod_cliente'],
            'tese_nulidade':tese_nulidade,
            'nr_processo': dados['nr_processo'],
            'data_publicacao':funcoes_gerais.convert_data_br(data_publicacao),
            'tipo_da_peca':'Embargos Omissao (somente com a Tese de Nulidade)'
            }
        return render(request, 'pecas/embargos_omissao.html', dados)


def v009_embargos_ultrapetita(request):
    return HttpResponse("<h1>Embargo Ultra-Petita</h1>")

def v010_embargos_contradicao(request):
    return HttpResponse("<h1>Embargo Contradição</h1>")



@login_required
def v004_peticoes(request):
    if (request.method == "POST"):
        pasta=request.POST['pasta']
        autor=request.POST['autor']
        tipo=request.POST["tipo"]
        nr_processo=request.POST['nr_processo']
        cliente=request.POST['cliente']
        comarca=request.POST['comarca']
        #conv_nome=request.POST['conveniado_nome']
        #conv_oab=request.POST['conveniado_oab']
        uf=request.POST['uf']
        juizo=request.POST['juizo']
        cod_cliente=request.POST['cod_cliente']
        #pasta_saj='0000'
        lista=funcoes_banco.view_pasta(pasta)
        pasta_saj = Pasta(lista[0],lista[1],lista[2],lista[3],lista[4],lista[5],lista[6],lista[7],lista[8],lista[9],lista[10],lista[11],lista[12],lista[13],lista[14],lista[15],lista[16],lista[17],lista[18])


        informacoesDaPasta=funcoes_banco.view_informacoesDaPasta(pasta)

        if tipo=='op_001_DesarquivamentoDevHonDupl':
            file = peticoes.modelos_peticoes('op_001_desarquivamentoDevHonDupl', dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_002_DevHPImprocedencia':
            file = peticoes.modelos_peticoes('op_002_devolucaoHPImprocedencia',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_003_DevHPImprocedenciaCDesarquiv':
            file = peticoes.modelos_peticoes('op_003_devolucaoHPImprocedenciaCDesarquiv',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_007_ReiterandoExpedicaoOficioDev':
             file = peticoes.modelos_peticoes('op_007_reiterandoExpedicaoOficioDev', dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_008_DispComprovanteTransf':
            file=peticoes.modelos_peticoes('op_008_dispComprovanteTransf',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_009_DispComprovanteTransfComDesarquiv':
            file=peticoes.modelos_peticoes('op_009_dispComprovanteTransfComDesarquiv',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_006_DevPernambuco':
            file=peticoes.modelos_peticoes('op_006_devolucaoPernambuco',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_004_DevHPExtintoSemResolucao':
            file=peticoes.modelos_peticoes('op_004_devHPExtintoSemResolucao',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_005_DevHPExtintoCDesarquivamento':
            file=peticoes.modelos_peticoes('op_005_devHPExtintoCDesarquivamento',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_010_JuntadaDePagamentoDaCondenacao':
            file=peticoes.modelos_peticoes('op_010_juntadaDePagamentoDaCondenacao',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_012_DevolucaoHPRN-1':
            file=peticoes.modelos_peticoes('op_012_devolucaoHPRN1',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_013_DevolucaoHPRN-2':
            file=peticoes.modelos_peticoes('op_013_devolucaoHPRN2',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_014_DevolucaoHPRN-3':
            file=peticoes.modelos_peticoes('op_014_devolucaoHPRN3',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_015_DevolucaoHPRN-4':
            file=peticoes.modelos_peticoes('op_015_devolucaoHPRN4',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_016_DevolucaoHPRN-5':
            file=peticoes.modelos_peticoes('op_016_devolucaoHPRN5',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_017_DevolucaoHPRN-6':
            file=peticoes.modelos_peticoes('op_017_devolucaoHPRN6',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_018_DevolucaoHPRN-7':
            file=peticoes.modelos_peticoes('op_018_devolucaoHPRN7',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_011_JuntadaDeCustasFinais':
            file=peticoes.modelos_peticoes('op_011_juntadaDeCustasFinais',dados_da_pasta=pasta_saj,info=informacoesDaPasta)
        elif tipo=='op_019_ImpugnAntecipDePericia':
            file=peticoes.modelos_peticoes('op_019_ImpugnAntecipDePericia',dados_da_pasta=pasta_saj,info=informacoesDaPasta)

        return redirect('pecas:download', docmto=file)
    else:
        dados = {
            'tipo_da_pecao': 'Peticões'
        }
        return render(request, 'pecas/peticoes.html', dados)


@login_required
def peticoes_teste(request):
    if (request.method == "POST"):
        pasta=request.POST['pasta']
        autor=request.POST['autor']
        tipo=request.POST["tipo"]
        nr_processo=request.POST['nr_processo']
        cliente=request.POST['cliente']
        comarca=request.POST['comarca']
        uf=request.POST['uf']
        juizo=request.POST['juizo']
        cod_cliente=request.POST['cod_cliente']
        conv_nome=request.POST['conveniado_nome']
        conv_oab=request.POST['conveniado_oab']
        if tipo=='Desarquivamento DHD':
            file=peticao_teste.fun_peticao_01(pasta,cod_cliente,autor,nr_processo,comarca,uf,cliente,juizo,conv_nome,conv_oab)
        return redirect('pecas:download', docmto=file)
    else:
        dados = {
            'tipo_da_pecao': 'Peticões Teste'
        }
        return render(request, 'pecas/peticoes_teste.html', dados)

def planilha(request):
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save(os.environ.get('DIR_DOCUMENTOS')+"/sample.xlsx")
    return HttpResponse("<h1>Gravando planilha</h1>")
